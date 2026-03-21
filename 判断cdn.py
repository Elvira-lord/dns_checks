import dns.resolver
from openpyxl import Workbook, load_workbook
import os

domain_list=[]
with open("domain.txt",'r',encoding='utf-8') as f:
    domain_list=[line.strip() for line in f.readlines() if line.strip()]

#创建表
excel_path="result.xlsx"
#判断文件是否存在
if os.path.exists(excel_path):
    wb=load_workbook(excel_path)
    ws=wb.active
    if ws["A1"].value!="域名":
        print("不存在表头---正在创建")
        ws["A1"] = "域名"
        ws["B1"] = "A记录列表"
        ws["C1"] = "CNAME"
        ws["D1"] = "CND查到的ip列表"
        ws["E1"] = "CDN检测注释"
        ws["F1"] = "真实IP"
        wb.save(excel_path)
        print("表头创建成功")
else:
    wb=Workbook()  #创建一个excel表对象
    ws=wb.active   #获取当前活动的工作表
    #设置表头
    ws["A1"]="域名"
    ws["B1"]="A记录列表"
    ws["C1"]="CNAME"
    ws["D1"]="CND查到的ip列表"
    ws["E1"]="CDN检测注释"
    ws["F1"]="真实IP"
    wb.save(excel_path)
    print("文件已经创建")



print("进行CDN检测")
last_row = 2  # 默认从第二行开始写入
for domain in domain_list:
    # 各个记录表
    A_List = []
    CNAME_List = []
    CND_Ip = []
    CDN_note = ''
    real_ip = '-'

    # print(domain_list)
    print("正在查询域名:", domain)
    dns_server = [
    # 国内 DNS
    '114.114.114.114',      # 114DNS
    '223.5.5.5',            # 阿里云 DNS
    '223.6.6.6',            # 阿里云 DNS 备用
    '119.29.29.29',         # 腾讯 DNSPod
    '180.76.76.76',         # 百度 DNS
    '1.2.4.8',              # CNNIC SDNS
    # 国外 DNS
    '8.8.8.8',              # Google DNS
    '8.8.4.4',              # Google DNS 备用
    '1.1.1.1',              # Cloudflare DNS
    ]

    # 解析域名
    def ip_check(doamin,zidingyi_Dns=None):
        #创建解析ip的解析器
        jieXiQi=dns.resolver.Resolver()
        if zidingyi_Dns:
            jieXiQi.nameservers=[zidingyi_Dns]    #这里只能传列表或者元组
        try:
            #看A记录
            A_jiLu=jieXiQi.resolve(domain, 'A')
            # print("A记录:",A_jiLu)
            A_List=[]
            for ip in A_jiLu:
                # print("IP:",ip.address)
                A_List.append(ip.address)
            return A_List
        except:
            # print(domain,">>无A记录")
            return []

    # 解析CNAME
    def cname_check(domain):
        try:
            cname=dns.resolver.resolve(domain, 'CNAME')
            # print("cname=",cname)
            cname_list=[]
            for name in cname:
                # print("CNAME:",name.target)
                cname_list.append(name.target)
            return cname_list
        except:
            # print(domain,">>无CNAME记录")
            return []


    A_list=ip_check(domain)
    if A_list:
        print("A记录列表:",A_list)
    else :
        print("无A记录")


    cname_list=cname_check(domain)
    if cname_list:
        print("CNAME列表:",cname_list)
    else:
        print("无CNAME记录")
    CNAME_List=cname_list

    ips=set()    #集合存，可以避免重复

    #用不同的DNS服务器进行解析
    for dns_ser in dns_server:
        dns_ip=ip_check(domain,dns_ser)
        # print(f"{dns_ser}>>>{dns_ip}")
        for ip in dns_ip:
            ips.add(ip)

    print(ips)
    CND_Ip=ips

    # cdn_keywords = ['cdn', 'cloud', 'tencent', 'aliyun', 'wangsu']
    cdn_keywords = [
    'cdn', 'cloud', 'cloudflare', 'akamai', 'fastly',
    'tencent', 'aliyun', 'alicdn', 'kunlun', 'wangsu',
    'qcloud', 'myqcloud', 'huaweicloud', 'baidubce',
    'qiniu', 'upai', 'chinanetcenter', 'ccgslb',
    'cloudfront', 'edgesuite', 'edgekey', 'cache', 'edge'
        ]
    for kw in cdn_keywords:
        flag=0
        for c in cname_list:
            if kw in c:
                print(f"{domain} 的cname检测到 {kw} !!!!!")
                flag=1
                break
        if flag==1:
            break

    if len(ips)>1:
        print(f"{domain}不同cdn存在多个ip!!!!!")
        CDN_note=f"{domain}不同cdn存在多个ip!!!!!"
    elif len(ips)==1:
        print(f"{domain}当前cdn查出唯一ip")
        CDN_note=f"{domain}当前cdn查出唯一ip"
        print(list(ips)[0])
        real_ip=list(ips)[0]
    else:
        print(f"{domain}未解析到IP")
        CDN_note = f"{domain}未解析到IP"
    print("-------------------------------------------------")
    A_List=A_list


    #==========写入excel=====================
    """
    域名: domain_list
    A记录列表: A_List
    CNAM记录列表: CNAME_List
    CDN查到的ip列表: CND_Ip
    CDN注释: CDN_note
    真实ip: real_ip
    """
    wb=load_workbook(excel_path)
    ws=wb.active
    while ws[f"A{last_row}"].value is not None:
        last_row+=1
    ws[f"A{last_row}"]=domain
    ws[f"B{last_row}"]=','.join(map(str, A_List)) if A_List else ''
    ws[f"C{last_row}"]=','.join(map(str, CNAME_List)) if CNAME_List else ''
    ws[f"D{last_row}"]=','.join(map(str, CND_Ip)) if CND_Ip else ''
    ws[f"E{last_row}"]=CDN_note
    ws[f"F{last_row}"]=real_ip
    wb.save(excel_path)

    while ws[f"A{last_row}"].value is not None:
        last_row+=1
    # print(last_row)





